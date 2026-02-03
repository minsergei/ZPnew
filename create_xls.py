from openpyxl import load_workbook, Workbook
from xls2xlsx import XLS2XLSX
from openpyxl.styles import Border, Side, Font, PatternFill


def save_and_format(wb, counter, service_number):
    ws = wb.active

    # Определяем стиль границ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    medium_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    # Жирный и 12 шрифт
    bold_font = Font(bold=True, size=12)

    # Проходим по всем заполненным ячейкам
    for row in ws.iter_rows():
        # Проверяем, есть ли слово "Дата" в какой-либо ячейке текущей строки
        is_date_row = any(
            isinstance(cell.value, str) and 'Дата' in cell.value
            for cell in row
        )
        is_period_row = any(
            isinstance(cell.value, str) and 'Период' in cell.value
            for cell in row
        )
        for cell in row:
            cell.border = thin_border
            if is_date_row or is_period_row:
                # Если нужно сохранить жирный шрифт:
                cell.font = bold_font
                cell.border = medium_border
                cell.fill = PatternFill(start_color='CCD7F0', end_color='CCD7F0', fill_type='solid')

            if isinstance(cell.value, (int, float)) and cell.value >= 500:
                cell.number_format = '#,##0 "₽"'

    # Автоподбор ширины колонок по тексту
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем букву колонки
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(f'calculations/{service_number}.xlsx')


def create_zp(path):
    x2x = XLS2XLSX(path)
    x2x.to_xlsx("xls_files/spreadsheet.xlsx")

    wb = load_workbook('xls_files/spreadsheet.xlsx')
    sheet = wb.active
    sheet.delete_cols(idx=2, amount=2)
    file_counter = 1
    new_wb = None
    new_ws = None

    for row in sheet.iter_rows(values_only=False):
        # Если в ячейке первого столбца текст "Атлас"
        if 'таб.№' in str(row[1].value):
            service_number = str(row[1].value)[-7:]
        if str(row[0].value).strip() == 'АО "НТЦ "Атлас"':
            # Сохраняем предыдущую книгу, если она была создана
            if new_wb:
                save_and_format(new_wb, file_counter, service_number)

            # Создаем новую книгу
            new_wb = Workbook()
            new_ws = new_wb.active

        # Если книга уже инициализирована, записываем строку
        if new_ws:
            new_ws.append([cell.value for cell in row])

    # Сохраняем последний файл
    if new_wb:
        save_and_format(new_wb, file_counter, service_number)
