import os
import openpyxl
import xlrd
from openpyxl.styles import Border, Side, Font


# открываем xls файл
def create_zp(path):
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    rows = [sheet.row_values(row, 0) for row in range(sheet.nrows)]

    # удаляем пустые записи
    for i in rows:
        while '' in i:
            i.remove('')
    # for i in rows:
    #     while ' ' in i:
    #         i.remove(' ')

    #определяем позиции начала расчетки
    list_nrows = []
    for i in range(len(rows)):
        if 'АО "НТЦ "Атлас"' in rows[i]:
            list_nrows.append(i)
    list_nrows.append(len(rows))


    # создаем расчетки для каждого сотрудника
    for i in range(len(list_nrows)-1):
        a = list_nrows[i]
        b = list_nrows[i+1]

        # print(rows[a:b])

        zp_list = rows[a:b]
        zp_name = zp_list[1][1][-7:] + '.xlsx'

        if os.path.exists(f"calculations/{zp_name}"):
            workbook = openpyxl.load_workbook(f"calculations/{zp_name}")
            sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
        # оформление бордюра
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for i in zp_list:
            if len(i) == 2:
                if 'Период' in i or 'Дата' in i:
                    i.insert(1, " ")
                    i.insert(1, " ")
                else:
                    i.insert(1, " ")
                    i.insert(1, " ")
                    i.insert(1, " ")
            if len(i) == 3:
                i.insert(1, " ")
                i.insert(1, " ")
            sheet.append(i)
            # Получаем номер последней добавленной строки
            current_row = sheet.max_row
            # Жирный курсив для нужных строк
            if 'Период' in i or 'Дата' in i:
                for cell in sheet[current_row]:
                    cell.font = Font(bold=True)
            # Применяем границы к каждой ячейке в новой строке
            for cell in sheet[current_row]:
                cell.border = thin_border

        for column_cells in sheet.columns:
            # Находим максимальную длину значения в столбце
            max_length = 0
            column_letter = column_cells[0].column_letter  # Получаем букву столбца (A, B, C...)

            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Устанавливаем ширину (добавляем небольшой запас)
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
        workbook.save(f"calculations/{zp_name}")
